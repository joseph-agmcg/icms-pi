"""Extração de dados de planilhas Excel de filiais (formato ICMS-PI: ATC, Normal, DIFAL). Inclui colunas ATC e DIF. ALIQUOTA."""

import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from icms_pi.logger import configurar_logger_da_aplicacao

logger = configurar_logger_da_aplicacao(__name__)

# Nome exato da coluna I.E. no Excel (estilo icms.xlsx). Outras variações aceitas abaixo.
NOME_COLUNA_IE = "INSC.ESTADUAL"

# Outros nomes aceitos para a coluna I.E. (planilhas em outros formatos)
SINONIMOS_IE = (
    "INSC.ESTADUAL",
    "INSC. ESTADUAL",
    "i.e.",
    "ie",
    "i..e.",
    "inscrição estadual",
    "inscricao estadual",
)

# Número máximo de linhas a varrer para encontrar o cabeçalho
MAX_LINHAS_BUSCA_CABECALHO = 25

# Linhas e colunas da área do título onde pode estar o período (ex.: 1/1/2026 ou jan-26)
MAX_LINHAS_BUSCA_PERIODO = 12
MAX_COLUNAS_BUSCA_PERIODO = 30

# Meses em português (formato jan-26, jan/26, etc.)
MESES_ABREV = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}

# Número de dígitos da I.E. no Piauí (preenchimento com zeros à esquerda quando faltar)
DIGITOS_IE_PI = 9


def _normalizar_cabecalho(texto: str | None) -> str:
    """Retorna o texto em minúsculo, sem acentos, sem pontos e com espaços normais."""
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = "".join(c for c in s if c != ".")
    return " ".join(s.split())


def _celula_bate_nome_ie(valor: str | None) -> bool:
    """Verifica se o valor da célula corresponde ao nome da coluna I.E. (ou sinônimos)."""
    n = _normalizar_cabecalho(valor)
    if not n:
        return False
    for nome in SINONIMOS_IE:
        if _normalizar_cabecalho(nome) == n:
            return True
    if n == "inscestadual" or n == "insc estadual":
        return True
    return False


def _encontrar_linha_cabecalho(
    planilha: openpyxl.worksheet.worksheet.Worksheet,
) -> int | None:
    """Retorna o índice 0-based da linha que contém a coluna I.E. (cabeçalho)."""
    for indice_linha in range(MAX_LINHAS_BUSCA_CABECALHO):
        row_num = indice_linha + 1
        linha = list(
            planilha.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )
        if not linha:
            continue
        celulas = linha[0]
        for valor in celulas:
            if _celula_bate_nome_ie(valor):
                return indice_linha
    return None


def _mapear_cabecalho(
    planilha: openpyxl.worksheet.worksheet.Worksheet, indice_linha: int
) -> tuple[dict[str, int], int]:
    """
    Lê a linha de cabeçalho e retorna (nome_coluna -> índice_coluna, índice_coluna_ie).
    Todas as colunas são mapeadas; colunas sem nome recebem "col_N". Chaves únicas.
    """
    row_num = indice_linha + 1
    linha = list(
        planilha.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
    )
    celulas = list(linha[0]) if linha else []
    nome_para_indice: dict[str, int] = {}
    coluna_ie: int | None = None
    for indice_col, valor in enumerate(celulas):
        nome = valor
        if nome is None or (isinstance(nome, str) and not nome.strip()):
            nome = f"col_{indice_col}"
        else:
            nome = str(nome).strip()
        if nome in nome_para_indice:
            nome = f"{nome}_{indice_col}"
        nome_para_indice[nome] = indice_col
        if _celula_bate_nome_ie(valor):
            coluna_ie = indice_col
    if coluna_ie is None:
        coluna_ie = -1
    return nome_para_indice, coluna_ie


def _extrair_periodo_da_area_titulo(
    planilha: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, int]:
    """
    Procura o período de referência na área do título da planilha (ex.: "APURAÇÃO DE ICMS PIAUI").
    Varre as primeiras linhas e colunas. Aceita:
    - Data: 1/1/2026, 01/01/2026 (dia/mês/ano ou mês/ano)
    - Texto com mês abreviado: jan-26, jan/26, jan-2026, jan/2026
    """
    for indice_linha in range(MAX_LINHAS_BUSCA_PERIODO):
        row_num = indice_linha + 1
        linha = list(
            planilha.iter_rows(
                min_row=row_num,
                max_row=row_num,
                min_col=1,
                max_col=MAX_COLUNAS_BUSCA_PERIODO,
                values_only=True,
            )
        )
        if not linha:
            continue
        celulas = linha[0]
        for valor in celulas:
            if not valor:
                continue
            if isinstance(valor, datetime):
                return valor.month, valor.year
            if isinstance(valor, (int, float)):
                continue
            texto = str(valor).strip().lower()
            mes_ano = _tentar_extrair_mes_ano_de_texto(texto)
            if mes_ano is not None:
                return mes_ano
    raise ValueError("Não foi possível identificar o período de referência na planilha.")


def _tentar_extrair_mes_ano_de_texto(texto: str) -> tuple[int, int] | None:
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.replace("\\", "/").replace("-", "/")

    padrao_data = re.compile(r"^(\\d{1,2})/(\\d{1,2})/(\\d{2,4})$")
    padrao_mes_ano = re.compile(r"^(\\d{1,2})/(\\d{2,4})$")

    m = padrao_data.match(texto)
    if m:
        dia, mes, ano = map(int, m.groups())
        if ano < 100:
            ano += 2000
        return mes, ano

    m = padrao_mes_ano.match(texto)
    if m:
        mes, ano = map(int, m.groups())
        if ano < 100:
            ano += 2000
        return mes, ano

    for abrev, mes in MESES_ABREV.items():
        if abrev in texto:
            numeros = re.findall(r"\\d{2,4}", texto)
            if numeros:
                ano = int(numeros[-1])
                if ano < 100:
                    ano += 2000
                return mes, ano
    return None


def _normalizar_ie_pi(ie: str | int | None) -> str:
    """Normaliza a I.E. do Piauí para um formato numérico fixo."""
    if ie is None:
        return ""
    s = str(ie).strip()
    s = "".join(c for c in s if c.isdigit())
    if not s:
        return ""
    return s.zfill(DIGITOS_IE_PI)


def _extrair_linhas_dados_completos(
    planilha: openpyxl.worksheet.worksheet.Worksheet,
    indice_linha_cabecalho: int,
    nome_para_indice: dict[str, int],
    coluna_ie: int,
) -> list[dict[str, object]]:
    """
    Percorre as linhas abaixo do cabeçalho e extrai todas as colunas até encontrar
    uma linha de rodapé / total ou fim da área de dados.
    """
    linhas: list[dict[str, object]] = []
    row_num = indice_linha_cabecalho + 2
    while True:
        linha = list(
            planilha.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )
        if not linha:
            break
        celulas = list(linha[0])
        if _linha_parece_rodape_ou_total(celulas):
            break
        dados: dict[str, object] = {}
        for nome_coluna, indice_coluna in nome_para_indice.items():
            if indice_coluna < len(celulas):
                dados[nome_coluna] = celulas[indice_coluna]
            else:
                dados[nome_coluna] = None
        if coluna_ie >= 0 and coluna_ie < len(celulas):
            dados["ie_normalizada"] = _normalizar_ie_pi(celulas[coluna_ie])
        else:
            dados["ie_normalizada"] = ""
        linhas.append(dados)
        row_num += 1
    return linhas


def _linha_parece_rodape_ou_total(celulas: list[object]) -> bool:
    """Heurística para detectar linhas de rodapé/total."""
    texto_concatenado = " ".join(str(c or "") for c in celulas).strip().lower()
    if not texto_concatenado:
        return True
    palavras_chave = (
        "total",
        "apresentacao ao fisco",
        "apresentação ao fisco",
        "sumario",
        "sumário",
        "fim",
        "piavai",
        "piaui",
    )
    return any(p in texto_concatenado for p in palavras_chave)


def extrair_todos_os_dados(
    caminho_arquivo: Path,
) -> tuple[list[dict[str, object]], dict[str, int], int, int]:
    """
    Carrega o arquivo Excel e retorna:
    - lista de dicionários com os dados de cada linha
    - mapeamento nome_coluna -> índice_coluna
    - mês de referência
    - ano de referência
    """
    logger.info("Carregando planilha: %s", caminho_arquivo)
    if not caminho_arquivo.exists():
        raise FileNotFoundError(caminho_arquivo)

    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    planilha = workbook.active

    indice_linha_cabecalho = _encontrar_linha_cabecalho(planilha)
    if indice_linha_cabecalho is None:
        raise ValueError("Não foi possível encontrar a linha de cabeçalho com I.E.")

    nome_para_indice, coluna_ie = _mapear_cabecalho(planilha, indice_linha_cabecalho)
    linhas = _extrair_linhas_dados_completos(
        planilha, indice_linha_cabecalho, nome_para_indice, coluna_ie
    )
    mes_ref, ano_ref = _extrair_periodo_da_area_titulo(planilha)

    logger.info(
        "Planilha carregada: %d linhas de dados, período %02d/%04d",
        len(linhas),
        mes_ref,
        ano_ref,
    )
    return linhas, nome_para_indice, mes_ref, ano_ref


def obter_dados_para_dae(
    linhas: list[dict[str, object]],
    nome_para_indice: dict[str, int],
    mes_ref: int,
    ano_ref: int,
) -> list[dict[str, object]]:
    """
    A partir da lista de linhas completas, retorna os dados necessários para a DAE:
    ie, ie_digitos, valor_atc, valor_normal (NORMAL), valor_difal (DIF. ALIQUOTA), mes_ref, ano_ref.
    """
    chave_ie = _obter_chave_ie(nome_para_indice)
    if chave_ie is None:
        raise ValueError("Não foi possível encontrar a coluna de I.E. na planilha.")

    lista: list[dict[str, object]] = []
    for dados in linhas:
        ie_norm = dados.get("ie_normalizada", "")
        valor_atc = _obter_valor_atc(dados)
        valor_normal = _obter_valor_normal(dados)
        valor_difal = _obter_valor_difal(dados)
        lista.append(
            {
                "ie": ie_norm,
                "ie_digitos": ie_norm,
                "valor_atc": valor_atc,
                "valor_normal": valor_normal,
                "valor_difal": valor_difal,
                "mes_ref": mes_ref,
                "ano_ref": ano_ref,
                "dados_originais": dados,
            }
        )
    logger.info("Total de registros para DAE: %d", len(lista))
    return lista


def _obter_chave_ie(nome_para_indice: dict[str, int]) -> str | None:
    """Encontra a chave do dicionário que representa a coluna de I.E."""
    for chave in nome_para_indice.keys():
        if _celula_bate_nome_ie(chave):
            return chave
    return None


def _obter_valor_atc(dados: dict[str, object]) -> float | None:
    """
    Retorna o valor da coluna ATC (valor principal) a partir dos dados de uma linha.
    Procura uma chave que contenha 'ATC' no nome.
    """
    for chave, valor in dados.items():
        if isinstance(chave, str) and "atc" in chave.lower():
            if isinstance(valor, (int, float)):
                return float(valor)
            if isinstance(valor, str):
                s = valor.strip().replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except ValueError:
                    continue
    return None


def _obter_valor_normal(dados: dict[str, object]) -> float | None:
    """
    Retorna o valor da coluna NORMAL a partir dos dados de uma linha.
    Procura chave cujo nome normalizado seja "normal" (ex.: "NORMAL"), sem incluir "ie_normalizada".
    """
    for chave, valor in dados.items():
        if not isinstance(chave, str):
            continue
        if chave.strip().lower() == "normal":
            if isinstance(valor, (int, float)):
                return float(valor)
            if isinstance(valor, str):
                s = valor.strip().replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except ValueError:
                    continue
    return None


def _obter_valor_difal(dados: dict[str, object]) -> float | None:
    """
    Retorna o valor da coluna DIF. ALIQUOTA (diferencial de alíquota) a partir dos dados de uma linha.
    Procura chave que contenha 'dif' e 'aliquota' no nome (ex.: "DIF. ALIQUOTA").
    """
    for chave, valor in dados.items():
        if not isinstance(chave, str):
            continue
        c = chave.lower()
        if "dif" in c and "aliquota" in c:
            if isinstance(valor, (int, float)):
                return float(valor)
            if isinstance(valor, str):
                s = valor.strip().replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except ValueError:
                    continue
    return None


def obter_ies_dos_dados(lista_dados: list[dict[str, object]]) -> list[str]:
    """Extrai apenas a lista de I.E.s normalizadas a partir da lista de dados."""
    ies: list[str] = []
    for item in lista_dados:
        ie = str(item.get("ie", "")).strip()
        if ie:
            ies.append(ie)
    return ies

